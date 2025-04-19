import sys
import ipaddress
import socket
import platform
import subprocess
import json
import os
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTextEdit, QMessageBox, QComboBox, QFileDialog, QTabWidget,
    QListWidget, QListWidgetItem, QInputDialog
)
from PyQt5.QtGui import QPalette, QColor, QFont
from PyQt5.QtCore import Qt, QSettings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class VLSMSubnettingApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🌐 Subnetting con VLSM - Herramienta Profesional")
        self.setGeometry(300, 100, 900, 700)
        self.settings = QSettings("RedesApp", "VLSMSubnetting")
        
        # Cargar configuración
        self.modo_oscuro = self.settings.value("modo_oscuro", False, type=bool)
        self.historial = json.loads(self.settings.value("historial", "[]"))
        self.ultima_ruta = self.settings.value("ultima_ruta", "")
        
        self.setup_ui()
        self.cargar_configuracion()
        
        # Variables de estado
        self.subredes_actuales = []
        self.calculo_actual = ""

    def setup_ui(self):
        # Configuración principal
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # Crear pestañas
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
        
        # Pestaña de cálculo
        self.tab_calculo = QWidget()
        self.tabs.addTab(self.tab_calculo, "🧮 Cálculo VLSM")
        self.setup_tab_calculo()
        
        # Pestaña de historial
        self.tab_historial = QWidget()
        self.tabs.addTab(self.tab_historial, "📚 Historial")
        self.setup_tab_historial()
        
        # Pestaña de herramientas
        self.tab_herramientas = QWidget()
        self.tabs.addTab(self.tab_herramientas, "🛠️ Herramientas")
        self.setup_tab_herramientas()
        
        # Barra de estado
        self.status_bar = QLabel("Listo")
        self.status_bar.setFont(QFont("Arial", 8))
        self.status_bar.setAlignment(Qt.AlignRight)
        layout.addWidget(self.status_bar)

    def setup_tab_calculo(self):
        layout = QVBoxLayout()
        self.tab_calculo.setLayout(layout)
        
        # Grupo de entrada de datos
        grupo_entrada = QVBoxLayout()
        
        self.label_ip = QLabel("IP de Red/CIDR (ej: 192.168.1.0/24 o 10.0.0.0/255.255.0.0):")
        self.input_ip = QLineEdit()
        self.input_ip.textChanged.connect(self.validar_ip)
        
        self.label_tipo = QLabel("Selecciona tipo de entrada:")
        self.combo_tipo = QComboBox()
        self.combo_tipo.addItems(["Cantidad de subredes", "Cantidad de hosts por subred"])
        self.combo_tipo.currentIndexChanged.connect(self.actualizar_placeholder)
        
        self.label_cantidad = QLabel("Valor:")
        self.input_cantidad = QLineEdit()
        self.input_cantidad.setPlaceholderText("Ej: 5 (subredes) o 30 (hosts)")
        
        grupo_entrada.addWidget(self.label_ip)
        grupo_entrada.addWidget(self.input_ip)
        grupo_entrada.addWidget(self.label_tipo)
        grupo_entrada.addWidget(self.combo_tipo)
        grupo_entrada.addWidget(self.label_cantidad)
        grupo_entrada.addWidget(self.input_cantidad)
        
        # Grupo de botones
        grupo_botones = QHBoxLayout()
        
        self.btn_calcular = QPushButton("🔍 Calcular Subnetting")
        self.btn_calcular.setToolTip("Realiza el cálculo de subnetting según los parámetros ingresados")
        self.btn_calcular.clicked.connect(self.calcular_subnetting)
        
        self.btn_limpiar = QPushButton("🧹 Limpiar")
        self.btn_limpiar.clicked.connect(self.limpiar_campos)
        
        grupo_botones.addWidget(self.btn_calcular)
        grupo_botones.addWidget(self.btn_limpiar)
        
        # Resultados
        self.resultado = QTextEdit()
        self.resultado.setReadOnly(True)
        self.resultado.setFont(QFont("Courier New", 10))
        
        # Grupo de exportación
        grupo_exportacion = QHBoxLayout()
        
        self.btn_exportar = QPushButton("📄 Exportar PDF")
        self.btn_exportar.clicked.connect(self.exportar_pdf)
        
        self.btn_exportar_excel = QPushButton("📊 Exportar a Excel")
        self.btn_exportar_excel.clicked.connect(self.exportar_excel)
        
        self.btn_grafico = QPushButton("📈 Ver gráfico")
        self.btn_grafico.clicked.connect(self.ver_grafico)
        
        grupo_exportacion.addWidget(self.btn_exportar)
        grupo_exportacion.addWidget(self.btn_exportar_excel)
        grupo_exportacion.addWidget(self.btn_grafico)
        
        # Ensamblar la pestaña
        layout.addLayout(grupo_entrada)
        layout.addLayout(grupo_botones)
        layout.addWidget(QLabel("📄 Resultados:"))
        layout.addWidget(self.resultado)
        layout.addLayout(grupo_exportacion)

    def setup_tab_historial(self):
        layout = QVBoxLayout()
        self.tab_historial.setLayout(layout)
        
        self.lista_historial = QListWidget()
        self.lista_historial.itemDoubleClicked.connect(self.cargar_desde_historial)
        
        self.btn_limpiar_historial = QPushButton("🧹 Limpiar Historial")
        self.btn_limpiar_historial.clicked.connect(self.limpiar_historial)
        
        layout.addWidget(QLabel("Historial de cálculos (doble click para cargar):"))
        layout.addWidget(self.lista_historial)
        layout.addWidget(self.btn_limpiar_historial)
        
        self.actualizar_lista_historial()

    def setup_tab_herramientas(self):
        layout = QVBoxLayout()
        self.tab_herramientas.setLayout(layout)
        
        # Herramientas de red
        grupo_red = QVBoxLayout()
        grupo_red.addWidget(QLabel("🛠️ Herramientas de Red:"))
        
        self.input_herramienta = QLineEdit()
        self.input_herramienta.setPlaceholderText("Ingrese IP o dominio")
        
        grupo_botones_red = QHBoxLayout()
        
        self.btn_ping = QPushButton("📶 Ping")
        self.btn_ping.clicked.connect(self.ping_ip)
        
        self.btn_resolver = QPushButton("🔎 Resolver DNS")
        self.btn_resolver.clicked.connect(self.resolver_ip)
        
        self.btn_escaneo = QPushButton("🔍 Escanear Puertos")
        self.btn_escaneo.clicked.connect(self.escanear_puertos)
        
        grupo_botones_red.addWidget(self.btn_ping)
        grupo_botones_red.addWidget(self.btn_resolver)
        grupo_botones_red.addWidget(self.btn_escaneo)
        
        grupo_red.addWidget(self.input_herramienta)
        grupo_red.addLayout(grupo_botones_red)
        
        # Herramientas de cálculo
        grupo_calculo = QVBoxLayout()
        grupo_calculo.addWidget(QLabel("🧮 Herramientas de Cálculo:"))
        
        self.btn_inverso = QPushButton("🔄 Cálculo inverso (IP a Red/CIDR)")
        self.btn_inverso.clicked.connect(self.calculo_inverso)
        
        self.btn_wildcard = QPushButton("🎭 Calcular Wildcard")
        self.btn_wildcard.clicked.connect(self.calcular_wildcard)
        
        grupo_calculo.addWidget(self.btn_inverso)
        grupo_calculo.addWidget(self.btn_wildcard)
        
        # Resultados de herramientas
        self.resultado_herramientas = QTextEdit()
        self.resultado_herramientas.setReadOnly(True)
        self.resultado_herramientas.setFont(QFont("Courier New", 10))
        
        # Modo visual
        grupo_modo = QHBoxLayout()
        self.btn_modo = QPushButton("🌙 Alternar Modo Claro/Oscuro")
        self.btn_modo.clicked.connect(self.alternar_modo)
        grupo_modo.addWidget(self.btn_modo)
        
        # Ensamblar la pestaña
        layout.addLayout(grupo_red)
        layout.addLayout(grupo_calculo)
        layout.addWidget(QLabel("📄 Resultados:"))
        layout.addWidget(self.resultado_herramientas)
        layout.addLayout(grupo_modo)

    def cargar_configuracion(self):
        # Aplicar modo oscuro si está activo
        self.aplicar_tema(self.modo_oscuro)
        
        # Cargar última IP usada
        ultima_ip = self.settings.value("ultima_ip", "")
        if ultima_ip:
            self.input_ip.setText(ultima_ip)

    def guardar_configuracion(self):
        self.settings.setValue("modo_oscuro", self.modo_oscuro)
        self.settings.setValue("historial", json.dumps(self.historial))
        self.settings.setValue("ultima_ruta", self.ultima_ruta)
        self.settings.setValue("ultima_ip", self.input_ip.text())

    def aplicar_tema(self, oscuro):
        palette = QPalette()
        if oscuro:
            palette.setColor(QPalette.Window, QColor(53, 53, 53))
            palette.setColor(QPalette.WindowText, Qt.white)
            palette.setColor(QPalette.Base, QColor(35, 35, 35))
            palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
            palette.setColor(QPalette.ToolTipBase, QColor(25, 25, 25))
            palette.setColor(QPalette.ToolTipText, Qt.white)
            palette.setColor(QPalette.Text, Qt.white)
            palette.setColor(QPalette.Button, QColor(53, 53, 53))
            palette.setColor(QPalette.ButtonText, Qt.white)
            palette.setColor(QPalette.BrightText, Qt.red)
            palette.setColor(QPalette.Highlight, QColor(142, 45, 197).lighter())
            palette.setColor(QPalette.HighlightedText, Qt.black)
            
            estilo = """
            QTextEdit, QListWidget {
                background-color: #232323;
                color: #ffffff;
                border: 1px solid #444;
            }
            QLineEdit {
                background-color: #353535;
                color: #ffffff;
                border: 1px solid #444;
            }
            """
        else:
            palette.setColor(QPalette.Window, QColor(240, 240, 240))
            palette.setColor(QPalette.WindowText, Qt.black)
            palette.setColor(QPalette.Base, Qt.white)
            palette.setColor(QPalette.AlternateBase, QColor(240, 240, 240))
            palette.setColor(QPalette.ToolTipBase, Qt.white)
            palette.setColor(QPalette.ToolTipText, Qt.black)
            palette.setColor(QPalette.Text, Qt.black)
            palette.setColor(QPalette.Button, QColor(240, 240, 240))
            palette.setColor(QPalette.ButtonText, Qt.black)
            palette.setColor(QPalette.BrightText, Qt.red)
            palette.setColor(QPalette.Highlight, QColor(100, 149, 237))
            palette.setColor(QPalette.HighlightedText, Qt.white)
            
            estilo = """
            QTextEdit, QListWidget {
                background-color: #ffffff;
                color: #000000;
                border: 1px solid #ccc;
            }
            QLineEdit {
                background-color: #ffffff;
                color: #000000;
                border: 1px solid #ccc;
            }
            """
        
        self.setPalette(palette)
        self.setStyleSheet(estilo)

    def actualizar_placeholder(self):
        if self.combo_tipo.currentText() == "Cantidad de subredes":
            self.input_cantidad.setPlaceholderText("Ej: 5 (número de subredes necesarias)")
        else:
            self.input_cantidad.setPlaceholderText("Ej: 30 (hosts por subred)")

    def validar_ip(self):
        texto = self.input_ip.text()
        try:
            ipaddress.IPv4Network(texto, strict=False)
            self.input_ip.setStyleSheet("background-color: #c8f7c5;")
            return True
        except ValueError:
            self.input_ip.setStyleSheet("background-color: #f7c5c5;")
            return False

    def calcular_subnetting(self):
        if not self.validar_ip():
            self.mostrar_error("❌ La dirección IP/CIDR no es válida.")
            return

        red_cidr = self.input_ip.text().strip()
        cantidad = self.input_cantidad.text().strip()
        modo = self.combo_tipo.currentText()

        if not cantidad:
            self.mostrar_error("❌ Por favor, ingresa un valor para el cálculo.")
            return

        try:
            cantidad = int(cantidad)
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            self.mostrar_error("❌ La cantidad debe ser un número entero positivo.")
            return

        try:
            red = ipaddress.IPv4Network(red_cidr, strict=False)
        except ValueError as e:
            self.mostrar_error(f"❌ Error en la red: {str(e)}")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        texto_resultado = f"📡 Subnetting de la red {red} ({timestamp}):\n\n"

        try:
            if modo == "Cantidad de subredes":
                if cantidad > (2 ** (30 - red.prefixlen)):
                    self.mostrar_error(f"❌ No se pueden crear {cantidad} subredes con una red /{red.prefixlen}.")
                    return

                bits_necesarios = (cantidad - 1).bit_length()
                nuevo_prefijo = red.prefixlen + bits_necesarios

                if nuevo_prefijo > 30:
                    self.mostrar_error("❌ No hay suficientes direcciones para crear tantas subredes.")
                    return

                subredes = list(red.subnets(new_prefix=nuevo_prefijo))
                texto_resultado += f"➡️ Subdivisión en {len(subredes)} subredes (/{nuevo_prefijo}):\n\n"

                for i, subred in enumerate(subredes[:cantidad]):
                    texto_resultado += self.info_subred(i + 1, subred)

            elif modo == "Cantidad de hosts por subred":
                if cantidad > (2 ** (32 - red.prefixlen - 1) - 2):
                    self.mostrar_error(f"❌ No se pueden crear subredes con {cantidad} hosts en una red /{red.prefixlen}.")
                    return

                bits_necesarios = (cantidad + 2 - 1).bit_length()
                nuevo_prefijo = 32 - bits_necesarios

                if nuevo_prefijo < red.prefixlen:
                    self.mostrar_error("❌ No hay suficiente espacio en la red original para subredes de ese tamaño.")
                    return

                subredes = list(red.subnets(new_prefix=nuevo_prefijo))
                texto_resultado += f"➡️ Subredes con al menos {cantidad} hosts (/{nuevo_prefijo}):\n\n"

                for i, subred in enumerate(subredes):
                    texto_resultado += self.info_subred(i + 1, subred)

            self.resultado.setPlainText(texto_resultado)
            self.subredes_actuales = subredes
            self.calculo_actual = texto_resultado

            # Guardar en historial
            entrada_historial = {
                "fecha": timestamp,
                "red": str(red),
                "modo": modo,
                "valor": cantidad,
                "resultado": texto_resultado
            }
            self.historial.append(entrada_historial)
            self.actualizar_lista_historial()
            self.guardar_configuracion()

            self.actualizar_status(f"Cálculo completado para {red}")

        except Exception as e:
            self.mostrar_error(f"❌ Error inesperado: {str(e)}")

    def info_subred(self, numero, subred):
        try:
            first_ip = subred.network_address + 1
            last_ip = subred.broadcast_address - 1
            num_hosts = subred.num_addresses - 2
        except Exception:
            first_ip = last_ip = "N/A"
            num_hosts = "N/A"

        texto = (
            f"Subred {numero}:\n"
            f"  ➤ Dirección de red: {subred.network_address}\n"
            f"  ➤ Broadcast: {subred.broadcast_address}\n"
            f"  ➤ Rango de IPs: {first_ip} - {last_ip}\n"
            f"  ➤ Hosts disponibles: {num_hosts}\n"
            f"  ➤ Máscara: {subred.netmask} (/{subred.prefixlen})\n"
            f"  ➤ Wildcard: {subred.hostmask}\n\n"
        )
        return texto

    def exportar_pdf(self):
        if not self.calculo_actual:
            self.mostrar_error("❌ No hay resultados para exportar.")
            return

        ruta, _ = QFileDialog.getSaveFileName(
            self, 
            "Guardar PDF", 
            self.ultima_ruta or "",
            "PDF (*.pdf)"
        )
        
        if not ruta:
            return

        try:
            # Actualizar última ruta
            self.ultima_ruta = os.path.dirname(ruta)
            self.guardar_configuracion()

            # Crear documento PDF profesional
            doc = SimpleDocTemplate(ruta, pagesize=letter)
            styles = getSampleStyleSheet()
            
            # Estilo personalizado para el título
            titulo_style = ParagraphStyle(
                'Titulo',
                parent=styles['Heading1'],
                fontSize=14,
                leading=18,
                spaceAfter=12,
                alignment=1  # Centrado
            )
            
            # Estilo para contenido
            contenido_style = ParagraphStyle(
                'Contenido',
                parent=styles['Normal'],
                fontName='Courier',
                fontSize=10,
                leading=12
            )
            
            # Contenido del PDF
            contenido = []
            
            # Título
            titulo = Paragraph("Reporte de Subnetting VLSM", titulo_style)
            contenido.append(titulo)
            contenido.append(Spacer(1, 12))
            
            # Fecha
            fecha = Paragraph(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
            contenido.append(fecha)
            contenido.append(Spacer(1, 12))
            
            # Agregar cada línea del resultado
            for linea in self.calculo_actual.split('\n'):
                if linea.strip():
                    p = Paragraph(linea.replace(' ', '&nbsp;'), contenido_style)
                    contenido.append(p)
                    contenido.append(Spacer(1, 6))
            
            # Generar PDF
            doc.build(contenido)
            
            self.actualizar_status(f"PDF exportado correctamente a {ruta}")
            QMessageBox.information(self, "Éxito", "✅ Reporte PDF generado correctamente.")
            
        except Exception as e:
            self.mostrar_error(f"❌ Error al exportar PDF: {str(e)}")

    def exportar_excel(self):
        if not hasattr(self, 'subredes_actuales') or not self.subredes_actuales:
            self.mostrar_error("❌ No hay subredes calculadas para exportar.")
            return

        ruta, _ = QFileDialog.getSaveFileName(
            self, 
            "Guardar Excel", 
            self.ultima_ruta or "",
            "Excel (*.xlsx)"
        )
        
        if not ruta:
            return

        try:
            # Actualizar última ruta
            self.ultima_ruta = os.path.dirname(ruta)
            self.guardar_configuracion()

            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Subnetting VLSM"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            center_alignment = Alignment(horizontal='center')
            
            # Encabezados
            encabezados = [
                "Subred", "Dirección de Red", "Primera IP", "Última IP", 
                "Broadcast", "Hosts", "Máscara", "Prefijo", "Wildcard"
            ]
            
            for col_num, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=1, column=col_num, value=encabezado)
                celda.font = header_font
                celda.fill = header_fill
                celda.border = border
                celda.alignment = center_alignment
            
            # Datos
            for row_num, subred in enumerate(self.subredes_actuales, 2):
                first_ip = subred.network_address + 1
                last_ip = subred.broadcast_address - 1
                num_hosts = subred.num_addresses - 2
                
                datos = [
                    f"Subred {row_num-1}",
                    str(subred.network_address),
                    str(first_ip),
                    str(last_ip),
                    str(subred.broadcast_address),
                    num_hosts,
                    str(subred.netmask),
                    f"/{subred.prefixlen}",
                    str(subred.hostmask)
                ]
                
                for col_num, dato in enumerate(datos, 1):
                    celda = ws.cell(row=row_num, column=col_num, value=dato)
                    celda.border = border
                    if col_num in (2, 3, 4, 5):  # Alinear direcciones IP al centro
                        celda.alignment = center_alignment
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            # Guardar archivo
            wb.save(ruta)
            
            self.actualizar_status(f"Excel exportado correctamente a {ruta}")
            QMessageBox.information(self, "Éxito", "✅ Archivo Excel generado correctamente.")
            
        except Exception as e:
            self.mostrar_error(f"❌ Error al exportar a Excel: {str(e)}")

    def ver_grafico(self):
        if not hasattr(self, 'subredes_actuales') or not self.subredes_actuales:
            self.mostrar_error("❌ Realiza un cálculo de subredes primero.")
            return
        
        try:
            # Preparar datos
            subredes = self.subredes_actuales[:20]  # Limitar para visualización
            nombres = [f"Subred {i+1}" for i in range(len(subredes))]
            hosts = [subred.num_addresses - 2 for subred in subredes]
            rangos = [f"{subred.network_address + 1}\n-\n{subred.broadcast_address - 1}" 
                     for subred in subredes]
            
            # Crear figura
            plt.figure(figsize=(12, 7))
            
            # Gráfico de barras
            bars = plt.bar(nombres, hosts, color='skyblue')
            
            # Añadir etiquetas y título
            plt.xlabel("Subredes")
            plt.ylabel("Número de Hosts")
            plt.title("Distribución de Hosts por Subred")
            plt.xticks(rotation=45, ha='right')
            
            # Añadir información adicional en las barras
            for bar, rango in zip(bars, rangos):
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2., height,
                         rango,
                         ha='center', va='bottom', rotation=0, fontsize=8)
            
            # Ajustar layout
            plt.tight_layout()
            
            # Mostrar gráfico
            plt.show()
            
        except Exception as e:
            self.mostrar_error(f"❌ Error al generar gráfico: {str(e)}")

    def ping_ip(self):
        objetivo = self.input_herramienta.text().strip()
        if not objetivo:
            self.mostrar_error("❌ Ingresa una dirección IP para hacer ping.", herramienta=True)
            return
        
        try:
            # Validar que sea una IP válida
            ipaddress.IPv4Address(objetivo)
            
            self.resultado_herramientas.setPlainText(f"🔍 Realizando ping a {objetivo}...")
            QApplication.processEvents()  # Actualizar la interfaz
            
            comando = ['ping', '-n' if platform.system() == 'Windows' else '-c', '4', objetivo]
            
            try:
                output = subprocess.check_output(
                    comando, 
                    stderr=subprocess.STDOUT, 
                    universal_newlines=True,
                    timeout=10
                )
                self.resultado_herramientas.setPlainText(f"📶 Resultado Ping a {objetivo}:\n{output}")
                self.actualizar_status(f"Ping completado a {objetivo}")
            except subprocess.TimeoutExpired:
                self.resultado_herramientas.setPlainText(f"⏳ Tiempo de espera agotado para {objetivo}")
            except subprocess.CalledProcessError as e:
                self.resultado_herramientas.setPlainText(f"❌ Error en ping a {objetivo}:\n{e.output}")
        
        except ValueError:
            self.mostrar_error(f"❌ {objetivo} no es una dirección IPv4 válida.", herramienta=True)

    def resolver_ip(self):
        objetivo = self.input_herramienta.text().strip()
        if not objetivo:
            self.mostrar_error("❌ Ingresa un dominio o IP para resolver.", herramienta=True)
            return
        
        try:
            self.resultado_herramientas.setPlainText(f"🔍 Resolviendo {objetivo}...")
            QApplication.processEvents()  # Actualizar la interfaz
            
            # Intentar resolver como dominio
            try:
                ip = socket.gethostbyname(objetivo)
                self.resultado_herramientas.append(f"🌐 {objetivo} resuelto a: {ip}")
                
                # Intentar resolver nombre si se ingresó una IP
                try:
                    nombre, _, _ = socket.gethostbyaddr(objetivo)
                    self.resultado_herramientas.append(f"🏷️ Nombre asociado: {nombre}")
                except:
                    pass
                
                self.actualizar_status(f"Resolución DNS completada para {objetivo}")
            
            except socket.gaierror:
                # Si falla, intentar como IP
                try:
                    ip_obj = ipaddress.IPv4Address(objetivo)
                    try:
                        nombre, _, _ = socket.gethostbyaddr(str(ip_obj))
                        self.resultado_herramientas.setPlainText(f"🏷️ {ip_obj} corresponde a: {nombre}")
                    except socket.herror:
                        self.resultado_herramientas.setPlainText(f"ℹ️ {ip_obj} es una IP válida pero no tiene nombre asociado")
                except ValueError:
                    self.mostrar_error(f"❌ No se pudo resolver {objetivo} como dominio ni como IP.", herramienta=True)
        
        except Exception as e:
            self.mostrar_error(f"❌ Error al resolver: {str(e)}", herramienta=True)

    def escanear_puertos(self):
        objetivo = self.input_herramienta.text().strip()
        if not objetivo:
            self.mostrar_error("❌ Ingresa una dirección IP para escanear.", herramienta=True)
            return
        
        try:
            ipaddress.IPv4Address(objetivo)
        except ValueError:
            self.mostrar_error(f"❌ {objetivo} no es una dirección IPv4 válida.", herramienta=True)
            return
        
        # Pedir rango de puertos
        inicio, ok = QInputDialog.getInt(
            self, "Escanear Puertos", "Puerto inicial:", 1, 1, 65535
        )
        if not ok:
            return
            
        fin, ok = QInputDialog.getInt(
            self, "Escanear Puertos", "Puerto final:", 100, inicio, 65535
        )
        if not ok:
            return
        
        self.resultado_herramientas.setPlainText(f"🔍 Escaneando puertos {inicio}-{fin} en {objetivo}...")
        QApplication.processEvents()  # Actualizar la interfaz
        
        puertos_abiertos = []
        
        for puerto in range(inicio, fin + 1):
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.settimeout(0.5)
                    resultado = s.connect_ex((objetivo, puerto))
                    if resultado == 0:
                        puertos_abiertos.append(puerto)
                        servicio = socket.getservbyport(puerto, 'tcp') if puerto <= 1024 else "?"
                        self.resultado_herramientas.append(f"✅ Puerto {puerto} ({servicio}) abierto")
                    else:
                        self.resultado_herramientas.append(f"❌ Puerto {puerto} cerrado")
                    QApplication.processEvents()
            except Exception as e:
                self.resultado_herramientas.append(f"⚠️ Error en puerto {puerto}: {str(e)}")
                continue
        
        if puertos_abiertos:
            self.resultado_herramientas.append("\n📌 Resumen de puertos abiertos:")
            for puerto in puertos_abiertos:
                servicio = socket.getservbyport(puerto, 'tcp') if puerto <= 1024 else "?"
                self.resultado_herramientas.append(f"  - {puerto} ({servicio})")
        
        self.actualizar_status(f"Escaneo completado para {objetivo} (puertos {inicio}-{fin})")

    def calculo_inverso(self):
        ip_str = self.input_herramienta.text().strip()
        if not ip_str:
            self.mostrar_error("❌ Ingresa una dirección IP para calcular.", herramienta=True)
            return
        
        try:
            ip = ipaddress.IPv4Address(ip_str)
            
            # Pedir máscara o prefijo
            mascara, ok = QInputDialog.getText(
                self, 
                "Cálculo Inverso", 
                "Ingresa máscara (ej: 255.255.255.0 o /24):"
            )
            if not ok or not mascara:
                return
            
            try:
                if mascara.startswith('/'):
                    red = ipaddress.IPv4Network(f"{ip_str}{mascara}", strict=False)
                else:
                    red = ipaddress.IPv4Network(f"{ip_str}/{mascara}", strict=False)
                
                resultado = (
                    f"🔄 Información de red para {ip_str} con máscara {mascara}:\n\n"
                    f"  ➤ Dirección de red: {red.network_address}\n"
                    f"  ➤ Broadcast: {red.broadcast_address}\n"
                    f"  ➤ Rango de hosts: {red.network_address + 1} - {red.broadcast_address - 1}\n"
                    f"  ➤ Máscara: {red.netmask} (/{red.prefixlen})\n"
                    f"  ➤ Wildcard: {red.hostmask}\n"
                    f"  ➤ Hosts disponibles: {red.num_addresses - 2}\n"
                )
                
                self.resultado_herramientas.setPlainText(resultado)
                self.actualizar_status(f"Cálculo inverso completado para {ip_str}")
                
            except ValueError as e:
                self.mostrar_error(f"❌ Máscara no válida: {str(e)}", herramienta=True)
        
        except ValueError:
            self.mostrar_error(f"❌ {ip_str} no es una dirección IPv4 válida.", herramienta=True)

    def calcular_wildcard(self):
        mascara = self.input_herramienta.text().strip()
        if not mascara:
            self.mostrar_error("❌ Ingresa una máscara para calcular su wildcard.", herramienta=True)
            return
        
        try:
            if mascara.startswith('/'):
                # Es un prefijo CIDR
                prefijo = int(mascara[1:])
                if not (0 <= prefijo <= 32):
                    raise ValueError
                mascara_obj = ipaddress.IPv4Network(f"0.0.0.0/{prefijo}").netmask
            else:
                # Es una máscara normal
                mascara_obj = ipaddress.IPv4Address(mascara)
            
            wildcard = ipaddress.IPv4Address(int(mascara_obj) ^ 0xFFFFFFFF)
            
            resultado = (
                f"🎭 Cálculo de Wildcard:\n\n"
                f"  ➤ Máscara ingresada: {mascara}\n"
                f"  ➤ Wildcard resultante: {wildcard}\n"
                f"  ➤ Prefijo CIDR equivalente: /{ipaddress.IPv4Network(f'0.0.0.0/{mascara}').prefixlen if not mascara.startswith('/') else mascara}\n"
            )
            
            self.resultado_herramientas.setPlainText(resultado)
            self.actualizar_status(f"Wildcard calculado para {mascara}")
        
        except ValueError:
            self.mostrar_error("❌ La máscara ingresada no es válida.", herramienta=True)

    def actualizar_lista_historial(self):
        self.lista_historial.clear()
        for item in reversed(self.historial):
            texto = f"{item['fecha']} - {item['red']} ({item['modo']}: {item['valor']})"
            QListWidgetItem(texto, self.lista_historial)

    def cargar_desde_historial(self, item):
        indice = len(self.historial) - self.lista_historial.row(item) - 1
        entrada = self.historial[indice]
        
        self.input_ip.setText(entrada["red"])
        self.combo_tipo.setCurrentText(entrada["modo"])
        self.input_cantidad.setText(str(entrada["valor"]))
        self.resultado.setPlainText(entrada["resultado"])
        
        # Intentar reconstruir las subredes para exportación
        try:
            red = ipaddress.IPv4Network(entrada["red"], strict=False)
            modo = entrada["modo"]
            valor = entrada["valor"]
            
            if modo == "Cantidad de subredes":
                bits_necesarios = (valor - 1).bit_length()
                nuevo_prefijo = red.prefixlen + bits_necesarios
                self.subredes_actuales = list(red.subnets(new_prefix=nuevo_prefijo))[:valor]
            else:
                bits_necesarios = (valor + 2 - 1).bit_length()
                nuevo_prefijo = 32 - bits_necesarios
                self.subredes_actuales = list(red.subnets(new_prefix=nuevo_prefijo))
            
            self.calculo_actual = entrada["resultado"]
            self.actualizar_status(f"Cálculo cargado desde historial: {entrada['fecha']}")
        except:
            self.subredes_actuales = []
            self.actualizar_status("Cálculo cargado desde historial (sin datos para exportar)")

    def limpiar_historial(self):
        respuesta = QMessageBox.question(
            self, 
            "Limpiar historial", 
            "¿Estás seguro de que quieres borrar todo el historial?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if respuesta == QMessageBox.Yes:
            self.historial = []
            self.actualizar_lista_historial()
            self.guardar_configuracion()
            self.actualizar_status("Historial limpiado")

    def limpiar_campos(self):
        self.input_ip.clear()
        self.input_cantidad.clear()
        self.resultado.clear()
        self.subredes_actuales = []
        self.calculo_actual = ""
        self.actualizar_status("Campos limpiados")

    def alternar_modo(self):
        self.modo_oscuro = not self.modo_oscuro
        self.aplicar_tema(self.modo_oscuro)
        self.guardar_configuracion()
        
        modo = "oscuro" if self.modo_oscuro else "claro"
        self.actualizar_status(f"Modo {modo} activado")

    def mostrar_error(self, mensaje, herramienta=False):
        QMessageBox.critical(self, "Error", mensaje)
        if herramienta:
            self.resultado_herramientas.append(mensaje)
        else:
            self.resultado.clear()

    def actualizar_status(self, mensaje):
        self.status_bar.setText(f"Estado: {mensaje}")

    def closeEvent(self, event):
        self.guardar_configuracion()
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Establecer estilo moderno
    app.setStyle('Fusion')
    
    ventana = VLSMSubnettingApp()
    ventana.show()
    sys.exit(app.exec_())